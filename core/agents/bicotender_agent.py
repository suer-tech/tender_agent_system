"""Агент глубокого анализа тендеров с Bicotender.

Полный пайплайн:
1. Поиск на bicotender.ru по ключевым словам
2. Переход в каждую карточку → чтение полного описания
3. Скачивание прикреплённых документов (PDF, DOCX, XLSX)
4. Парсинг документов → извлечение текста
5. Отправка контекста в Claude → глубокий анализ релевантности
6. Формирование сводного отчёта
"""
from __future__ import annotations

import json
import re
import subprocess
import textwrap
from pathlib import Path

import sys
ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(ROOT))

from core.sources import bicotender


# --------------- парсеры документов ---------------

def _read_pdf(path: Path) -> str:
    """Извлечь текст из PDF через pdfplumber."""
    try:
        import pdfplumber
        text_parts: list[str] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages[:30]:  # макс 30 страниц
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return "\n".join(text_parts)
    except ImportError:
        return "[pdfplumber не установлен — pip install pdfplumber]"
    except Exception as e:
        return f"[ошибка чтения PDF: {e}]"


def _read_docx(path: Path) -> str:
    """Извлечь текст из DOCX через python-docx."""
    try:
        from docx import Document
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return "[python-docx не установлен — pip install python-docx]"
    except Exception as e:
        return f"[ошибка чтения DOCX: {e}]"


def _read_doc(path: Path) -> str:
    """Извлечь текст из .doc (старый Word формат).

    Стратегии (по приоритету):
    1. olefile — читает OLE-структуру напрямую (чистый Python)
    2. Переименование в .docx и попытка чтения через python-docx
    3. Чтение сырых байт и извлечение текста регуляркой
    """
    # Стратегия 1: olefile
    try:
        import olefile
        ole = olefile.OleFileIO(str(path))
        if ole.exists("WordDocument"):
            # Пробуем извлечь текст из потока
            streams = [s for s in ole.listdir() if "text" in str(s).lower() or "word" in str(s).lower()]
            parts = []
            for stream_path in ole.listdir():
                try:
                    data = ole.openstream(stream_path).read()
                    text = data.decode("utf-8", errors="replace")
                    # Фильтруем мусор — оставляем только строки с буквами
                    clean = [l for l in text.splitlines() if sum(c.isalpha() for c in l) > len(l) * 0.3]
                    if clean:
                        parts.extend(clean)
                except Exception:
                    pass
            ole.close()
            if parts:
                return "\n".join(parts)[:10000]
        ole.close()
    except ImportError:
        pass
    except Exception:
        pass

    # Стратегия 2: грубое извлечение текста из бинарника
    try:
        raw = path.read_bytes()
        # .doc хранит текст в ASCII/CP1251 — пробуем обе кодировки
        for enc in ("cp1251", "utf-8", "latin-1"):
            try:
                text = raw.decode(enc, errors="replace")
                # Выделяем фрагменты из >= 4 подряд идущих букв/пробелов
                fragments = re.findall(r"[а-яА-ЯёЁa-zA-Z0-9\s.,;:!?\-()«»\"]{20,}", text)
                if fragments:
                    result = "\n".join(fragments)
                    # Убираем дубли и мусор
                    result = re.sub(r"\s{3,}", "\n", result)
                    return result[:10000]
            except Exception:
                continue
    except Exception:
        pass

    return "[.doc — не удалось извлечь текст]"


def _read_xlsx(path: Path) -> str:
    """Извлечь текст из XLSX через openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        parts: list[str] = []
        for ws in wb.worksheets[:5]:
            parts.append(f"--- Лист: {ws.title} ---")
            for row in ws.iter_rows(max_row=200, values_only=True):
                vals = [str(v) for v in row if v is not None]
                if vals:
                    parts.append(" | ".join(vals))
        return "\n".join(parts)
    except ImportError:
        return "[openpyxl не установлен — pip install openpyxl]"
    except Exception as e:
        return f"[ошибка чтения XLSX: {e}]"


def _read_xls(path: Path) -> str:
    """Извлечь текст из старого .xls формата через xlrd."""
    try:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        parts: list[str] = []
        for ws in wb.sheets()[:5]:
            parts.append(f"--- Лист: {ws.name} ---")
            for row_idx in range(min(ws.nrows, 200)):
                vals = [str(ws.cell_value(row_idx, col)) for col in range(ws.ncols)
                        if ws.cell_value(row_idx, col)]
                if vals:
                    parts.append(" | ".join(vals))
        return "\n".join(parts)
    except ImportError:
        # Fallback: пробуем openpyxl (иногда .xls файлы на самом деле xlsx)
        try:
            return _read_xlsx(path)
        except Exception:
            return "[xlrd не установлен — pip install xlrd]"
    except Exception as e:
        return f"[ошибка чтения XLS: {e}]"


def _read_html(path: Path) -> str:
    """Извлечь текст из HTML."""
    try:
        # Пробуем разные кодировки
        raw = None
        for enc in ("utf-8", "cp1251", "latin-1"):
            try:
                raw = path.read_text(encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        if not raw:
            raw = path.read_text(encoding="utf-8", errors="replace")
        text = re.sub(r"<script[^>]*>.*?</script>", "", raw, flags=re.DOTALL | re.I)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL | re.I)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return text[:10000]
    except Exception as e:
        return f"[ошибка чтения HTML: {e}]"


def _read_rtf(path: Path) -> str:
    """Извлечь текст из RTF."""
    try:
        from striprtf.striprtf import rtf_to_text
        raw = path.read_text(encoding="utf-8", errors="replace")
        return rtf_to_text(raw)[:10000]
    except ImportError:
        # Грубый fallback — убираем RTF-теги
        try:
            raw = path.read_text(encoding="utf-8", errors="replace")
            text = re.sub(r"\\[a-z]+\d*\s?", " ", raw)
            text = re.sub(r"[{}]", "", text)
            text = re.sub(r"\s+", " ", text).strip()
            return text[:10000] if len(text) > 50 else "[RTF: слишком мало текста]"
        except Exception:
            return "[RTF: ошибка чтения]"


def _unpack_archive(path: Path) -> list[Path]:
    """Распаковать ZIP/RAR архив во временную папку. Возвращает список файлов."""
    import zipfile
    import tempfile

    out_dir = path.parent / f"{path.stem}_unpacked"
    out_dir.mkdir(exist_ok=True)

    extracted: list[Path] = []

    if zipfile.is_zipfile(path):
        try:
            with zipfile.ZipFile(path, 'r') as zf:
                for info in zf.infolist()[:20]:  # макс 20 файлов
                    if info.is_dir():
                        continue
                    # Фикс кодировки имён (cp437 → cp866 для русских имён)
                    try:
                        fname = info.filename.encode('cp437').decode('cp866')
                    except (UnicodeDecodeError, UnicodeEncodeError):
                        fname = info.filename
                    target = out_dir / Path(fname).name
                    with zf.open(info) as src, open(target, 'wb') as dst:
                        dst.write(src.read())
                    extracted.append(target)
        except Exception as e:
            print(f"[doc_parser] ZIP error: {e}")
    else:
        # Пробуем RAR через rarfile
        try:
            import rarfile
            with rarfile.RarFile(str(path)) as rf:
                for info in rf.infolist()[:20]:
                    if info.is_dir():
                        continue
                    target = out_dir / Path(info.filename).name
                    with rf.open(info) as src, open(target, 'wb') as dst:
                        dst.write(src.read())
                    extracted.append(target)
        except ImportError:
            print("[doc_parser] rarfile не установлен для .rar")
        except Exception as e:
            print(f"[doc_parser] RAR error: {e}")

    return extracted


def parse_document(path: Path) -> str:
    """Универсальный парсер: выбирает нужный ридер по расширению.

    Поддерживаемые форматы:
    - PDF (.pdf) — pdfplumber
    - Word (.docx) — python-docx
    - Word старый (.doc) — olefile / сырое извлечение
    - Excel (.xlsx) — openpyxl
    - Excel старый (.xls) — xlrd / openpyxl fallback
    - HTML (.html, .htm) — regex strip tags
    - RTF (.rtf) — striprtf / regex fallback
    - Текст (.txt, .csv) — прямое чтение
    - Архивы (.zip, .rar) — распаковка + парсинг содержимого
    """
    ext = path.suffix.lower()

    # Архивы — распаковываем и парсим содержимое
    if ext in (".zip", ".rar"):
        files = _unpack_archive(path)
        if not files:
            return f"[архив {ext}: не удалось распаковать]"
        parts: list[str] = []
        for f in files:
            text = parse_document(f)  # рекурсия
            if text and not text.startswith("["):
                parts.append(f"--- {f.name} ---\n{text}")
        return "\n\n".join(parts) if parts else f"[архив {ext}: нет читаемых файлов]"

    if ext == ".pdf":
        return _read_pdf(path)
    elif ext == ".docx":
        return _read_docx(path)
    elif ext == ".doc":
        return _read_doc(path)
    elif ext == ".xlsx":
        return _read_xlsx(path)
    elif ext == ".xls":
        return _read_xls(path)
    elif ext in (".html", ".htm"):
        return _read_html(path)
    elif ext == ".rtf":
        return _read_rtf(path)
    elif ext in (".txt", ".csv"):
        try:
            for enc in ("utf-8", "cp1251", "latin-1"):
                try:
                    return path.read_text(encoding=enc)[:10000]
                except UnicodeDecodeError:
                    continue
            return path.read_text(encoding="utf-8", errors="replace")[:10000]
        except Exception as e:
            return f"[ошибка чтения {ext}: {e}]"
    else:
        return f"[неподдерживаемый формат: {ext}]"


# --------------- анализ через Claude ---------------

ANALYSIS_PROMPT = textwrap.dedent("""\
    Ты — эксперт-аналитик тендеров. Проанализируй тендер ниже.
    У тебя есть описание со страницы тендера и содержимое прикреплённых документов.

    Запрос пользователя: {user_query}

    Оцени, насколько этот тендер соответствует запросу пользователя.
    Учитывай тематику, отрасль, требования и специфику запроса.
    Проведи ГЛУБОКИЙ анализ.

    Ответь СТРОГО валидным JSON:
    {{
        "relevant": true|false,
        "score": 0-10,
        "summary": "2-3 предложения — суть закупки",
        "key_requirements": ["требование 1", "требование 2", ...],
        "budget": "бюджет если указан",
        "deadline_info": "сроки подачи и исполнения",
        "qualification": "требования к участнику (опыт, лицензии, допуски)",
        "tech_stack": "упомянутые технологии, платформы, стандарты",
        "risks": "потенциальные риски участия",
        "recommendation": "конкретная рекомендация — участвовать или нет, и почему",
        "reason": "краткое обоснование оценки релевантности"
    }}

    === ТЕНДЕР ===
    Название: {title}
    Заказчик: {customer}
    Цена: {price}
    Дедлайн: {deadline}
    URL: {url}

    === ОПИСАНИЕ СО СТРАНИЦЫ ===
    {detail_text}

    === ДОКУМЕНТЫ ===
    {documents_text}
""")


def _call_claude(prompt: str, timeout: int = 180) -> dict | None:
    """Вызов Claude CLI для анализа."""
    try:
        res = subprocess.run(
            ["claude", "-p", prompt],
            capture_output=True, text=True, timeout=timeout, encoding="utf-8",
        )
    except (subprocess.TimeoutExpired, FileNotFoundError) as e:
        return {"error": str(e)}

    out = (res.stdout or "").strip()
    match = re.search(r"\{.*\}", out, re.DOTALL)
    if not match:
        return {"error": "no_json", "raw": out[:500]}
    try:
        return json.loads(match.group(0))
    except json.JSONDecodeError:
        return {"error": "bad_json", "raw": out[:500]}


def analyze_tender(tender: dict, user_query: str, timeout: int = 180) -> dict:
    """Глубокий анализ одного тендера через Claude."""
    # Собираем тексты документов
    doc_texts: list[str] = []
    for doc_path in tender.get("doc_files", []):
        p = Path(doc_path) if not isinstance(doc_path, Path) else doc_path
        if p.exists():
            text = parse_document(p)
            if text and not text.startswith("["):
                doc_texts.append(f"--- {p.name} ---\n{text[:5000]}")

    if doc_texts:
        documents_text = "\n\n".join(doc_texts)
    else:
        doc_files = tender.get("doc_files", [])
        if not doc_files:
            documents_text = "(документы: не найдены на странице тендера или не удалось скачать)"
        else:
            documents_text = "(документы: скачаны, но не удалось извлечь текст)"

    detail_text = tender.get("detail_text", "") or tender.get("description", "")

    prompt = ANALYSIS_PROMPT.format(
        user_query=user_query,
        title=tender.get("title", ""),
        customer=tender.get("customer", ""),
        price=tender.get("price", ""),
        deadline=tender.get("deadline", ""),
        url=tender.get("url", ""),
        detail_text=detail_text[:6000],
        documents_text=documents_text[:8000],
    )

    return _call_claude(prompt, timeout=timeout)


# --------------- сводный отчёт ---------------

SUMMARY_PROMPT = textwrap.dedent("""\
    Ты — эксперт-аналитик тендеров. Сформируй СВОДНЫЙ ОТЧЁТ по результатам анализа.

    Запрос пользователя: {user_query}

    Ниже — JSON-массив с результатами анализа каждого тендера.
    Сформируй краткий отчёт на русском языке:

    1. ОБЩАЯ СВОДКА: сколько найдено, сколько релевантных, общие тренды
    2. ТОП РЕКОМЕНДАЦИИ: самые перспективные тендеры (score >= 7)
    3. ПО КАЖДОМУ РЕЛЕВАНТНОМУ ТЕНДЕРУ:
       - Название и заказчик
       - Суть закупки
       - Бюджет и сроки
       - Ключевые требования
       - Твоя рекомендация
    4. ТЕНДЕРЫ С ПОТЕНЦИАЛОМ (score 4-6) — кратко
    5. ИТОГОВАЯ РЕКОМЕНДАЦИЯ

    Отвечай в формате текста с markdown-разметкой. НЕ JSON.

    === ДАННЫЕ ===
    {data}
""")


def generate_report(results: list[dict], user_query: str) -> str:
    """Генерирует сводный отчёт по всем проанализированным тендерам."""
    # Готовим данные для сводки
    summary_data = []
    for r in results:
        analysis = r.get("analysis", {})
        summary_data.append({
            "title": r.get("title", ""),
            "customer": r.get("customer", ""),
            "price": r.get("price", ""),
            "deadline": r.get("deadline", ""),
            "url": r.get("url", ""),
            "score": analysis.get("score", 0),
            "relevant": analysis.get("relevant", False),
            "summary": analysis.get("summary", ""),
            "key_requirements": analysis.get("key_requirements", []),
            "budget": analysis.get("budget", ""),
            "qualification": analysis.get("qualification", ""),
            "tech_stack": analysis.get("tech_stack", ""),
            "recommendation": analysis.get("recommendation", ""),
            "reason": analysis.get("reason", ""),
            "risks": analysis.get("risks", ""),
        })

    data_json = json.dumps(summary_data, ensure_ascii=False, indent=2)
    # Обрезаем если слишком много
    if len(data_json) > 15000:
        data_json = data_json[:15000] + "\n... (обрезано)"

    prompt = SUMMARY_PROMPT.format(user_query=user_query, data=data_json)

    try:
        res = subprocess.run(
            ["claude", "-p", prompt],
            capture_output=True, text=True, timeout=300, encoding="utf-8",
        )
        return (res.stdout or "").strip()
    except Exception as e:
        return f"Ошибка генерации отчёта: {e}"


# --------------- основной пайплайн ---------------

def run(
    user_query: str,
    keywords: list[str] | None = None,
    limit: int = 10,
    headless: bool = True,
    max_docs: int = 5,
    progress_callback=None,
) -> dict:
    """Полный пайплайн агента.

    Args:
        user_query: описание того, что ищет пользователь
        keywords: список ключевых слов для поиска (если None — берёт из config.yaml)
        limit: макс. тендеров на один ключ
        headless: запуск браузера без GUI
        max_docs: макс. документов на тендер
        progress_callback: callable(msg: str) для уведомлений о прогрессе

    Returns:
        {
            "query": str,
            "total_found": int,
            "analyzed": int,
            "relevant_count": int,
            "results": [тендер + analysis],
            "report": str,
        }
    """
    import yaml

    def _log(msg: str):
        print(msg)
        if progress_callback:
            progress_callback(msg)

    if not keywords:
        with open(ROOT / "config.yaml", "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        keywords = cfg.get("bicotender", {}).get("keywords", cfg.get("keywords", []))

    all_tenders: list[dict] = []
    seen_ids: set[str] = set()

    _log(f"Начинаю поиск по {len(keywords)} ключевым словам...")

    for i, kw in enumerate(keywords, 1):
        _log(f"[{i}/{len(keywords)}] Поиск: «{kw}»")
        try:
            tenders = bicotender.search_and_collect(
                keyword=kw,
                limit=limit,
                headless=headless,
                max_docs_per_tender=max_docs,
            )
            for t in tenders:
                eid = t.get("external_id", "")
                if eid not in seen_ids:
                    seen_ids.add(eid)
                    all_tenders.append(t)
            _log(f"  найдено: {len(tenders)}, новых: {len(tenders) - (len(tenders) - len([t for t in tenders if t.get('external_id') in seen_ids]))}")
        except Exception as e:
            _log(f"  ошибка: {e}")

    _log(f"\nВсего уникальных тендеров: {len(all_tenders)}. Начинаю анализ через Claude...")

    # Анализ каждого тендера
    for i, tender in enumerate(all_tenders, 1):
        _log(f"[{i}/{len(all_tenders)}] Анализ: {tender.get('title', '')[:60]}...")
        analysis = analyze_tender(tender, user_query)
        tender["analysis"] = analysis
        score = analysis.get("score", 0) if not analysis.get("error") else 0
        _log(f"  оценка: {score}/10")

    relevant = [t for t in all_tenders if t.get("analysis", {}).get("relevant")]
    _log(f"\nРелевантных: {len(relevant)} из {len(all_tenders)}. Генерирую отчёт...")

    report = generate_report(all_tenders, user_query)

    return {
        "query": user_query,
        "total_found": len(all_tenders),
        "analyzed": len(all_tenders),
        "relevant_count": len(relevant),
        "results": all_tenders,
        "report": report,
    }
